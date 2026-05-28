#!/usr/bin/env python3
"""Extract selected metrics from the Robin Nature supplementary tables workbook."""

from __future__ import annotations

import argparse
import json
from collections import defaultdict
from statistics import mean
from typing import Any

from openpyxl import load_workbook


DRUGS_OF_INTEREST = ("DMSO", "Y-27632", "Ripasudil", "MFGE8")


def _round(value: float) -> float:
    return round(value, 4)


def summarize_mfi(workbook: Any) -> dict[str, dict[str, Any]]:
    summaries: dict[str, dict[str, Any]] = {}
    for sheet in ("Supp Table 3", "Supp Table 4", "Supp Table 5"):
        ws = workbook[sheet]
        rows = list(ws.iter_rows(values_only=True))[1:]
        values: dict[str, list[float]] = defaultdict(list)
        for row in rows:
            drug = row[0]
            norm = row[4]
            if drug in DRUGS_OF_INTEREST and isinstance(norm, (int, float)):
                values[str(drug)].append(float(norm))
        summaries[sheet] = {
            drug: {
                "n": len(vals),
                "mean_normalized_mfi": _round(mean(vals)),
                "values": [_round(v) for v in vals],
            }
            for drug, vals in values.items()
        }
    return summaries


def summarize_bixbench(workbook: Any) -> dict[str, Any]:
    ws = workbook["Supp Table 10"]
    rows = list(ws.iter_rows(values_only=True))
    header = rows[0]
    idx = {name: i for i, name in enumerate(header)}

    def aggregate(group_col: str | None) -> dict[str, dict[str, Any]]:
        data: dict[str, list[float]] = defaultdict(lambda: [0.0, 0.0, 0.0])
        for row in rows[1:]:
            key = "overall" if group_col is None else str(row[idx[group_col]])
            data[key][0] += float(row[idx["finch_accuracy"]] or 0)
            data[key][1] += float(row[idx["llm_accuracy"]] or 0)
            data[key][2] += 1
        return {
            key: {
                "n": int(vals[2]),
                "finch_mean_accuracy": _round(vals[0] / vals[2]),
                "llm_mean_accuracy": _round(vals[1] / vals[2]),
            }
            for key, vals in sorted(data.items())
        }

    return {
        "overall": aggregate(None)["overall"],
        "by_category": aggregate("category"),
        "by_original_category": aggregate("original_category"),
    }


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("xlsx", help="Path to Nature supplementary tables xlsx.")
    parser.add_argument("--json", action="store_true", help="Emit JSON.")
    args = parser.parse_args()

    workbook = load_workbook(args.xlsx, read_only=True, data_only=True)
    summary = {
        "mfi": summarize_mfi(workbook),
        "bixbench": summarize_bixbench(workbook),
    }

    if args.json:
        print(json.dumps(summary, indent=2, ensure_ascii=False))
        return

    print("MFI means")
    for sheet, drugs in summary["mfi"].items():
        print(f"\n{sheet}")
        for drug, vals in drugs.items():
            print(f"- {drug}: n={vals['n']}, mean={vals['mean_normalized_mfi']}")

    overall = summary["bixbench"]["overall"]
    print("\nBixBench-style table")
    print(
        f"- overall: n={overall['n']}, Finch={overall['finch_mean_accuracy']}, "
        f"LLM={overall['llm_mean_accuracy']}"
    )
    for key, vals in summary["bixbench"]["by_category"].items():
        print(
            f"- {key}: n={vals['n']}, Finch={vals['finch_mean_accuracy']}, "
            f"LLM={vals['llm_mean_accuracy']}"
        )


if __name__ == "__main__":
    main()
